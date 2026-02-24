"""
Two-Week Scheduler
==================

Reads an existing two-week schedule template Excel file, identifies every
unassigned D/N cell across the 14 working days, and assigns specific shift
codes using seniority + preference + rest-rule logic.

Column layout is detected dynamically from the header row, so it works
across different sheet versions of the template.
"""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import io

# ---------------------------------------------------------------------------
# Shift metadata
# ---------------------------------------------------------------------------

SHIFT_TIMES: Dict[str, Tuple[int, int, int, int]] = {
    "D7B":   (7,  0, 19,  0),
    "D7P":   (7,  0, 19,  0),
    "D9L":   (9,  0, 21,  0),
    "D11M":  (11, 0, 23,  0),
    "D11H":  (11, 0, 23,  0),
    "MG":    (11, 0, 23,  0),
    "LG":    (9,  0, 21,  0),
    "PG":    (7,  0, 19,  0),
    "GR":    (7,  0, 19,  0),
    "FLOAT": (9,  0, 21,  0),
    "N7B":   (19, 0,  7,  0),
    "N7P":   (19, 0,  7,  0),
    "N9L":   (21, 0,  9,  0),
    "NG":    (19, 0,  7,  0),
    "NP":    (19, 0,  7,  0),
}

DAY_SHIFT_CODES   = {"D7B","D7P","D9L","D11M","D11H","MG","LG","PG","GR","FLOAT"}
NIGHT_SHIFT_CODES = {"N7B","N7P","N9L","NG","NP"}
ALL_SHIFT_CODES   = DAY_SHIFT_CODES | NIGHT_SHIFT_CODES

RW_SHIFTS       = {"D7B","D9L","D11M","D11H","N7B"}
GR_SHIFTS       = {"MG","GR","LG","PG","NP"}
COMBINED_SHIFTS = {"D7P","N7P","N9L","NG"}

ADMIN_CODES = {"AT", "LT-D", "LT-N"}

# Values/patterns that mean the person is absent that day
ABSENT_KEYWORDS = {
    "LOA","MIL","SM","EDU","SIM","STABLE","ATLS","CLINICAL",
    "OFF","AOC","KCU","COMM","FLOAT",
}

BASE_TO_DAY_SHIFTS: Dict[str, List[str]] = {
    "B":  ["D7B","GR"],
    "H":  ["D11H"],
    "L":  ["D9L","LG"],
    "P":  ["D7P","PG"],
    "M":  ["D11M","MG"],
    "B7": ["D7B","GR"],
    "B11":["D11H"],
}
BASE_TO_NIGHT_SHIFTS: Dict[str, List[str]] = {
    "B": ["N7B","NG"],
    "L": ["N9L"],
    "P": ["N7P","NP"],
}

DAY_SHIFT_PRIORITY   = ["D7B","D7P","D9L","D11M","D11H","MG","LG","PG","GR"]
NIGHT_SHIFT_PRIORITY = ["N7B","N7P","N9L","NG","NP"]

STANDARD_REST = 12
REDUCED_REST  = 10

HEADER_SEARCH_ROWS = range(30, 50)
FIRST_STAFF_OFFSET  = 2  # staff starts this many rows after header


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _str(val) -> str:
    if val is None: return ""
    return str(val).strip()

def _int(v) -> int:
    if v is None: return 0
    try: return int(float(str(v)))
    except: return 0

def _float(v) -> float:
    if v is None: return 0.0
    try: return float(str(v))
    except: return 0.0

def _norm(val: str) -> str:
    """Strip trailing 'p' (dual-as-medic marker) and uppercase."""
    return re.sub(r'p$', '', val.upper())

def is_absent(val: str) -> bool:
    v = val.upper()
    if v.startswith("^"): return True
    for kw in ABSENT_KEYWORDS:
        if kw in v: return True
    if re.search(r'\b(EDU|SIM|STABLE|ATLS|CLINICAL|AOC|LOA|MIL|SM|OFF|MIL)\b', v):
        return True
    return False

def is_already_assigned(val: str) -> bool:
    return _norm(val) in ALL_SHIFT_CODES

def is_admin(val: str) -> bool:
    v = val.upper()
    return v in {"AT"} or v.startswith("LT-")

def counts_as_shift(val: str) -> bool:
    if not val: return False
    return is_already_assigned(val) or is_admin(val)


# ---------------------------------------------------------------------------
# Shift time helpers
# ---------------------------------------------------------------------------

def _shift_end(code: str, date: datetime) -> Optional[datetime]:
    c = _norm(code)
    if c not in SHIFT_TIMES: return None
    sh, sm, eh, em = SHIFT_TIMES[c]
    s = date.replace(hour=sh, minute=sm, second=0, microsecond=0)
    e = date.replace(hour=eh, minute=em, second=0, microsecond=0)
    if e <= s: e += timedelta(days=1)
    return e

def _shift_start(code: str, date: datetime) -> Optional[datetime]:
    c = _norm(code)
    if c not in SHIFT_TIMES: return None
    sh, sm, _, _ = SHIFT_TIMES[c]
    return date.replace(hour=sh, minute=sm, second=0, microsecond=0)

def rest_hours(prev_code: str, prev_date: datetime,
               next_code: str, next_date: datetime) -> float:
    end   = _shift_end(prev_code, prev_date)
    start = _shift_start(next_code, next_date)
    if end is None or start is None: return 99.0
    return (start - end).total_seconds() / 3600


# ---------------------------------------------------------------------------
# Column layout detector
# ---------------------------------------------------------------------------

class SheetLayout:
    """
    Detects the column layout of a schedule sheet dynamically by reading
    the header row. Stores 1-based column indices.
    """
    def __init__(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        self.ws = ws
        self.header_row: int = 0
        self.date_cols: List[int]      = []   # 1-based col indices of all date cols (prev + 14 working)
        self.date_dates: List[datetime] = []  # datetime for each date col
        self.col_role: int    = 1
        self.col_matrix: int  = 2
        self.col_dual: int    = 4
        self.col_seniority: int = 7
        self.col_name: int    = 9   # will be overridden
        self.col_wk1_tgt: int = 11
        self.col_wk2_tgt: int = 12
        self.col_wk1_cnt: int = 13
        self.col_wk2_cnt: int = 14
        self.col_flex: int    = 46
        self.col_10hr: int    = 47
        self.col_rw: int      = 42
        self.col_gr: int      = 43
        # Day base pref cols: key=base_label, val=col
        self.day_pref_cols: Dict[str, int]   = {}
        self.night_pref_cols: Dict[str, int] = {}

        self._detect()

    def _detect(self):
        ws = self.ws
        for row_idx in HEADER_SEARCH_ROWS:
            row_vals = {c: ws.cell(row=row_idx, column=c).value for c in range(1, 55)}
            # Look for the row that has 14+ datetime values in date-range columns
            date_found = [(c, v) for c, v in row_vals.items() if isinstance(v, datetime)]
            if len(date_found) >= 14:
                self.header_row = row_idx
                # All date columns, sorted
                date_found.sort(key=lambda x: x[0])
                for col, dt in date_found:
                    self.date_cols.append(col)
                    self.date_dates.append(dt)

                # Detect name column: find first row below header with 'n' or 'm' in col A
                # and find which column has a string name (not a number/date)
                for r in range(row_idx + 1, row_idx + 10):
                    role_val = _str(ws.cell(row=r, column=1).value).lower()
                    if role_val in ('n', 'm'):
                        # Name is the first column after DOH that has a string
                        for c in range(5, 20):
                            v = ws.cell(row=r, column=c).value
                            if isinstance(v, str) and len(v) > 1 and not v.lower() in ('na','nb','ma','mb','d','x'):
                                self.col_name = c
                                break
                        break

                # Wk1/Wk2 target and count: find 'Wk1', 'Wk2' labels
                for c, v in row_vals.items():
                    if _str(v) == 'Wk1':
                        self.col_wk1_tgt = c
                        self.col_wk1_cnt = c + 1  # counter right after
                    elif _str(v) == 'Wk2':
                        self.col_wk2_tgt = c
                        self.col_wk2_cnt = c + 1

                # Seniority: look for 'Sr' or 'M' label
                # col G (7) in 29-Mar, col G (7) in 1-Mar but it's 'M' (roster number)
                # Both have seniority as col 7; roster number differs
                self.col_seniority = 7

                # D/N Flex and 10hr Turn: search by label
                for c, v in row_vals.items():
                    sv = _str(v)
                    if 'flex' in sv.lower() or sv == 'D/N Flex':
                        self.col_flex = c
                    elif '10h' in sv.lower():
                        self.col_10hr = c
                    elif sv == 'RW':
                        self.col_rw = c
                    elif sv == 'GR':
                        self.col_gr = c

                # Day preference cols: look for B/H/L/P/M labels in the pref area
                # They appear twice: once for day, once for night
                label_cols: List[Tuple[int,str]] = []
                for c, v in sorted(row_vals.items()):
                    sv = _str(v).upper()
                    if sv in ('B','H','L','P','M','B7','B11'):
                        label_cols.append((c, sv))

                # First group = day prefs, second = night prefs
                if label_cols:
                    # Day pref cols are before night pref cols
                    # Night pref cols are a subset (B, L, P only)
                    # Split: first contiguous group = day
                    prev_c = None
                    day_group = []
                    night_group = []
                    in_night = False
                    for c, lbl in label_cols:
                        if prev_c and c - prev_c > 3:
                            in_night = True
                        if in_night:
                            night_group.append((c, lbl))
                        else:
                            day_group.append((c, lbl))
                        prev_c = c
                    for c, lbl in day_group:
                        self.day_pref_cols[lbl] = c
                    for c, lbl in night_group:
                        self.night_pref_cols[lbl] = c

                break

        if not self.header_row:
            raise ValueError("Could not find date header row in the schedule sheet.")

    @property
    def first_staff_row(self) -> int:
        return self.header_row + FIRST_STAFF_OFFSET

    @property
    def prev_date_cols(self) -> List[int]:
        """The two reference columns from the prior period."""
        return self.date_cols[:2]

    @property
    def working_date_cols(self) -> List[int]:
        """The 14 working day columns."""
        return self.date_cols[2:16]

    @property
    def prev_dates(self) -> List[datetime]:
        return self.date_dates[:2]

    @property
    def working_dates(self) -> List[datetime]:
        return self.date_dates[2:16]

    def all_date_col_set(self) -> set:
        return set(self.date_cols)


# ---------------------------------------------------------------------------
# Staff member
# ---------------------------------------------------------------------------

class StaffMember:
    def __init__(self, ws, row_idx: int, layout: SheetLayout):
        def get(col):
            return ws.cell(row=row_idx, column=col).value

        self.row_idx    = row_idx
        self.role       = _str(get(layout.col_role)).lower()
        self.matrix     = _str(get(layout.col_matrix)).lower()
        col4            = _str(get(4)).lower()
        self.is_dual    = col4 == 'd'
        self.seniority  = get(layout.col_seniority)
        self.name       = _str(get(layout.col_name))
        self.wk1_target = _int(get(layout.col_wk1_tgt))
        self.wk2_target = _int(get(layout.col_wk2_tgt))
        self.wk1_count  = _int(get(layout.col_wk1_cnt))
        self.wk2_count  = _int(get(layout.col_wk2_cnt))
        self.flex       = _str(get(layout.col_flex))
        self.ten_hr     = _str(get(layout.col_10hr)).lower() == 'yes'
        self.rw_count   = _float(get(layout.col_rw))
        self.gr_count   = _float(get(layout.col_gr))
        self.no_matrix  = 1 if self.matrix == 'a' else 0
        self.is_open_matrix = (self.matrix == 'a')

        # Day/night base preferences
        self.day_prefs: Dict[str, int] = {}
        for base, col in layout.day_pref_cols.items():
            v = get(col)
            if v is not None and _str(v) not in ('', 'None'):
                try: self.day_prefs[base] = int(float(_str(v)))
                except: pass

        self.night_prefs: Dict[str, int] = {}
        for base, col in layout.night_pref_cols.items():
            v = get(col)
            if v is not None and _str(v) not in ('', 'None'):
                try: self.night_prefs[base] = int(float(_str(v)))
                except: pass

        # Cell values for all date columns (prev + 14 working)
        # cells[0], cells[1] = prior period ref days
        # cells[2]..cells[15] = 14 working days
        all_cols = layout.date_cols  # list of 1-based col indices
        self.cells: List[str] = []
        for col in all_cols[:16]:
            self.cells.append(_str(get(col)))

        self.dates: List[datetime] = layout.date_dates[:16]

    # --- convenience accessors ---

    def cell(self, day_idx: int) -> str:
        """day_idx 0-13 = working days 1-14."""
        return self.cells[day_idx + 2]

    def set_cell(self, day_idx: int, val: str):
        self.cells[day_idx + 2] = val

    def prior_cell(self, day_idx: int) -> str:
        """Cell immediately before working day day_idx."""
        return self.cells[day_idx + 1]

    def date(self, day_idx: int) -> datetime:
        return self.dates[day_idx + 2]

    def prior_date(self, day_idx: int) -> datetime:
        return self.dates[day_idx + 1]

    def week(self, day_idx: int) -> int:
        return 1 if day_idx < 7 else 2

    def shifts_remaining(self, day_idx: int) -> int:
        w = self.week(day_idx)
        return max(0, (self.wk1_target if w == 1 else self.wk2_target)
                    - (self.wk1_count  if w == 1 else self.wk2_count))

    def add_shift(self, day_idx: int, shift_code: str):
        """Record an assignment: update cell and week counter."""
        self.set_cell(day_idx, shift_code)
        if self.week(day_idx) == 1:
            self.wk1_count += 1
        else:
            self.wk2_count += 1
        # Update RW/GR running totals
        code = _norm(shift_code)
        if code in RW_SHIFTS:
            self.rw_count += 1
        elif code in GR_SHIFTS:
            self.gr_count += 1
        elif code in COMBINED_SHIFTS:
            self.rw_count += 0.5
            self.gr_count += 0.5

    def preferred_shifts(self, is_day: bool) -> List[Tuple[str, int]]:
        """Return (shift_code, rank) pairs sorted by rank ascending."""
        prefs = self.day_prefs if is_day else self.night_prefs
        base_map = BASE_TO_DAY_SHIFTS if is_day else BASE_TO_NIGHT_SHIFTS
        result = []
        for base, rank in sorted(prefs.items(), key=lambda x: x[1]):
            for shift in base_map.get(base, []):
                result.append((shift, rank))
        return result

    def rw_gr_pref(self) -> str:
        if abs(self.rw_count - self.gr_count) < 0.5: return 'either'
        return 'GR' if self.rw_count > self.gr_count else 'RW'

    @property
    def seniority_key(self):
        try: return float(str(self.seniority))
        except: return 9999.0


# ---------------------------------------------------------------------------
# Eligibility checks
# ---------------------------------------------------------------------------

def _check_rest(staff: StaffMember, day_idx: int, shift_code: str) -> Tuple[bool, str]:
    prior = staff.prior_cell(day_idx)
    if not prior or is_absent(prior) or is_admin(prior): return True, ""
    pc = _norm(prior)
    if pc not in SHIFT_TIMES: return True, ""
    hours = rest_hours(pc, staff.prior_date(day_idx), shift_code, staff.date(day_idx))
    needed = REDUCED_REST if staff.ten_hr else STANDARD_REST
    if hours < needed:
        return False, f"Only {hours:.1f}h rest (need {needed}h)"
    return True, ""


def _check_consecutive(staff: StaffMember, day_idx: int, is_day: bool) -> Tuple[bool, str]:
    """Check the 4-in-a-row / 5-max rules."""
    # Build backward sequence of shifts (up to 5 back)
    seq = []
    for back in range(1, 6):
        ci = day_idx + 2 - back  # index into cells[]
        if ci < 0: break
        val = staff.cells[ci]
        if not val or is_absent(val) or is_admin(val): break
        code = _norm(val)
        if code in ALL_SHIFT_CODES:
            seq.append(code)
        else:
            break

    if not seq: return True, ""

    # Night-to-day is never allowed
    if is_day and seq[0] in NIGHT_SHIFT_CODES:
        return False, "Prior shift was a night (N→D not allowed)"

    # Count same-type streak
    same = 0
    for code in seq:
        if is_day and code in DAY_SHIFT_CODES: same += 1
        elif not is_day and code in NIGHT_SHIFT_CODES: same += 1
        else: break
    if same >= 4:
        return False, f"Already {same} consecutive {'day' if is_day else 'night'} shifts"

    # Total stretch including this assignment
    total = len(seq) + 1
    if total > 5:
        return False, "Would exceed 5-shift stretch limit"
    if total == 5:
        # Need a D→N transition somewhere in the stretch
        all_codes = [_norm(staff.cells[day_idx + 2 - b])
                     for b in range(1, 5) if (day_idx + 2 - b) >= 0]
        has_dn = False
        for i in range(len(all_codes) - 1):
            if all_codes[i] in DAY_SHIFT_CODES and all_codes[i+1] in NIGHT_SHIFT_CODES:
                has_dn = True; break
        # Also: if assigning a night and prior was a day → that IS the transition
        if not is_day and seq and seq[0] in DAY_SHIFT_CODES:
            has_dn = True
        if not has_dn:
            return False, "5th shift requires a Day→Night transition"

    return True, ""


def can_work(staff: StaffMember, day_idx: int, shift_code: str, is_day: bool) -> Tuple[bool, str]:
    ok, r = _check_rest(staff, day_idx, shift_code)
    if not ok: return False, r
    ok, r = _check_consecutive(staff, day_idx, is_day)
    if not ok: return False, r
    return True, ""


# ---------------------------------------------------------------------------
# Template reader
# ---------------------------------------------------------------------------

def read_template(wb: openpyxl.Workbook, sheet_name: str
                  ) -> Tuple[List[StaffMember], SheetLayout]:
    ws     = wb[sheet_name]
    layout = SheetLayout(ws)

    staff_list: List[StaffMember] = []
    for row_idx in range(layout.first_staff_row, ws.max_row + 1):
        role = _str(ws.cell(row=row_idx, column=layout.col_role).value).lower()
        if role not in ('n', 'm'):
            if staff_list and row_idx > layout.first_staff_row + len(staff_list) + 5:
                break
            continue
        name = _str(ws.cell(row=row_idx, column=layout.col_name).value)
        if not name or name in {'Senior RN','Junior RN','Senior Medic','Junior Medic','SUM','Counters Start'}:
            break
        # Skip rows where name looks like a number or is clearly a counter row
        if re.match(r'^\d+$', name):
            continue
        member = StaffMember(ws, row_idx, layout)
        if member.name:
            staff_list.append(member)

    return staff_list, layout


# ---------------------------------------------------------------------------
# Core assignment for one day
# ---------------------------------------------------------------------------

def _open_matrix_satisfied(slot_assignments: Dict[str, Optional[str]],
                            staff_by_name: Dict[str, StaffMember],
                            shift_code: str) -> bool:
    name = slot_assignments.get(shift_code)
    if name is None: return False
    s = staff_by_name.get(name)
    return s is not None and s.is_open_matrix


def assign_one_period(
    staff: List[StaffMember],
    day_idx: int,
    is_day: bool,
    pre_assigned: Dict[str, str],         # name → shift_code (already in cells)
    staff_by_name: Dict[str, StaffMember],
) -> Dict[str, Optional[str]]:
    """
    Assign all available D (or N) staff on day_idx to specific shift codes.
    Returns { shift_code: assigned_name_or_None }
    """
    target_cell = 'D' if is_day else 'N'
    shift_list  = DAY_SHIFT_PRIORITY if is_day else NIGHT_SHIFT_PRIORITY

    # Collect pre-assigned slots
    slots: Dict[str, Optional[str]] = {}
    assigned_names: set = set()

    for s in staff:
        cell = s.cell(day_idx)
        code = _norm(cell)
        if is_day and code in DAY_SHIFT_CODES:
            slots[code] = s.name
            assigned_names.add(s.name)
        elif not is_day and code in NIGHT_SHIFT_CODES:
            slots[code] = s.name
            assigned_names.add(s.name)

    # Collect unassigned-but-available staff
    candidates: List[StaffMember] = []
    for s in staff:
        if s.name in assigned_names: continue
        cell = s.cell(day_idx)
        if cell == target_cell:
            if s.shifts_remaining(day_idx) > 0:
                candidates.append(s)
        elif is_day and cell == 'N' and s.flex.lower() == 'yes':
            # Flex eligible: prior must not be a night
            prior = _norm(s.prior_cell(day_idx))
            if prior not in NIGHT_SHIFT_CODES and s.shifts_remaining(day_idx) > 0:
                candidates.append(s)

    # Sort by seniority
    candidates.sort(key=lambda s: s.seniority_key)

    # Open enough shift slots for available candidates
    open_needed = len(candidates)
    for sh in shift_list:
        if sh not in slots and open_needed > 0:
            slots[sh] = None
            open_needed -= 1

    # Track which slots still need an open-matrix person
    needs_open_matrix = {
        sh for sh in slots
        if not _open_matrix_satisfied(slots, staff_by_name, sh)
    }

    # Assign each candidate to best eligible shift
    for s in candidates:
        if s.name in assigned_names: continue
        if s.shifts_remaining(day_idx) <= 0: continue

        open_slots = [sh for sh in slots if slots[sh] is None]
        if not open_slots: break

        # Build ordered preference list
        pref_pairs = s.preferred_shifts(is_day)
        pref_ordered = [sh for sh, _ in pref_pairs if sh in open_slots]
        for sh in shift_list:
            if sh in open_slots and sh not in pref_ordered:
                pref_ordered.append(sh)

        # Re-order by RW/GR preference (soft)
        rw_gr = s.rw_gr_pref()
        if rw_gr == 'RW':
            pref_ordered.sort(key=lambda x: (0 if x in RW_SHIFTS else 1 if x in COMBINED_SHIFTS else 2))
        elif rw_gr == 'GR':
            pref_ordered.sort(key=lambda x: (0 if x in GR_SHIFTS else 1 if x in COMBINED_SHIFTS else 2))

        for sh in pref_ordered:
            if slots.get(sh) is not None: continue  # taken

            # Open-matrix check: if slot needs open-matrix and person isn't, skip
            # unless no open-matrix candidates remain (fall back)
            if sh in needs_open_matrix and not s.is_open_matrix:
                # Check if any remaining open-matrix candidate can fill this slot
                om_can_fill = any(
                    r.is_open_matrix and r.name not in assigned_names
                    and r.shifts_remaining(day_idx) > 0
                    for r in candidates[candidates.index(s)+1:]
                    if can_work(r, day_idx, sh, is_day)[0]
                )
                if om_can_fill:
                    continue  # defer — let an open-matrix person take it

            ok, _ = can_work(s, day_idx, sh, is_day)
            if not ok: continue

            slots[sh] = s.name
            assigned_names.add(s.name)
            s.add_shift(day_idx, sh)
            if sh in needs_open_matrix and s.is_open_matrix:
                needs_open_matrix.discard(sh)
            break

    return slots


# ---------------------------------------------------------------------------
# Main two-week loop
# ---------------------------------------------------------------------------

def run_schedule(wb: openpyxl.Workbook, sheet_name: str
                 ) -> Tuple[Dict, List[StaffMember], SheetLayout]:
    staff, layout = read_template(wb, sheet_name)
    staff_by_name = {s.name: s for s in staff}

    results: Dict[int, Dict] = {}
    for day_idx in range(14):
        date = layout.working_dates[day_idx]

        day_slots   = assign_one_period(staff, day_idx, True,  {}, staff_by_name)
        night_slots = assign_one_period(staff, day_idx, False, {}, staff_by_name)

        results[day_idx] = {
            'date':  date,
            'day':   day_slots,
            'night': night_slots,
        }

    return results, staff, layout


# ---------------------------------------------------------------------------
# Output workbook
# ---------------------------------------------------------------------------

def build_output(
    source_wb: openpyxl.Workbook,
    sheet_name: str,
    results: Dict,
    staff: List[StaffMember],
    layout: SheetLayout,
) -> openpyxl.Workbook:
    source_ws = source_wb[sheet_name]
    out_wb    = openpyxl.Workbook()
    out_ws    = out_wb.active
    out_ws.title = f"ASSIGNED"

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    BLUE   = PatternFill("solid", fgColor="DDEBF7")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")

    # Copy all rows up to and including header + 1
    for r in range(1, layout.first_staff_row):
        for c in range(1, source_ws.max_column + 1):
            v = source_ws.cell(row=r, column=c).value
            out_ws.cell(row=r, column=c).value = v

    # Build reverse lookup: (day_idx, name) → shift_code
    assignment_lookup: Dict[Tuple[int,str], str] = {}
    for day_idx, day_data in results.items():
        for sh, name in day_data['day'].items():
            if name: assignment_lookup[(day_idx, name)] = sh
        for sh, name in day_data['night'].items():
            if name: assignment_lookup[(day_idx, name)] = sh

    # Find source row for each staff member
    name_to_src_row = {}
    for r in range(layout.first_staff_row, source_ws.max_row + 1):
        nm = _str(source_ws.cell(row=r, column=layout.col_name).value)
        if nm: name_to_src_row[nm] = r

    # Write staff rows
    out_row = layout.first_staff_row
    for s in staff:
        src_r = name_to_src_row.get(s.name)
        if src_r is None: continue

        # Copy all metadata columns
        for c in range(1, min(layout.date_cols[0], source_ws.max_column + 1)):
            out_ws.cell(row=out_row, column=c).value = source_ws.cell(row=src_r, column=c).value

        out_ws.cell(row=out_row, column=layout.col_name).font = Font(bold=True)

        # Date columns
        for col_i, col_num in enumerate(layout.date_cols[:16]):
            orig_val = _str(source_ws.cell(row=src_r, column=col_num).value)
            dst = out_ws.cell(row=out_row, column=col_num)

            if col_i < 2:
                # Prior period reference — copy as-is
                dst.value = orig_val
                continue

            day_idx = col_i - 2  # 0-13

            if is_already_assigned(_norm(orig_val)) or is_admin(orig_val) or is_absent(orig_val) or orig_val.startswith('^') or not orig_val:
                dst.value = orig_val
                if is_already_assigned(_norm(orig_val)):
                    dst.fill = BLUE
            elif orig_val in ('D', 'N'):
                assigned = assignment_lookup.get((day_idx, s.name))
                if assigned:
                    dst.value = assigned
                    dst.fill  = GREEN
                else:
                    dst.value = 'OPEN'
                    dst.fill  = RED
            else:
                dst.value = orig_val

        out_row += 1

    # Column widths
    out_ws.column_dimensions[chr(64 + layout.col_name)].width = 18
    date_start_col = layout.date_cols[2] if len(layout.date_cols) > 2 else 17
    for col_num in range(date_start_col, date_start_col + 14):
        if 1 <= col_num <= 26:
            letter = chr(64 + col_num)
        else:
            letter = chr(64 + (col_num - 1) // 26) + chr(64 + (col_num - 1) % 26 + 1)
        out_ws.column_dimensions[letter].width = 9

    return out_wb


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def _summary(results: Dict, staff: List[StaffMember]) -> Dict:
    total = filled = open_ = 0
    days = []
    for day_idx in range(14):
        d = results.get(day_idx, {})
        date_obj = d.get('date')
        try: ds = date_obj.strftime('%-d %b %a')
        except: ds = f'Day {day_idx+1}'
        day_s   = d.get('day',   {})
        night_s = d.get('night', {})
        df = sum(1 for v in day_s.values()   if v)
        do = sum(1 for v in day_s.values()   if not v)
        nf = sum(1 for v in night_s.values() if v)
        no = sum(1 for v in night_s.values() if not v)
        total  += df + do + nf + no
        filled += df + nf
        open_  += do + no
        days.append({'date': ds, 'day_filled': df, 'day_open': do,
                     'night_filled': nf, 'night_open': no,
                     'day_shifts': day_s, 'night_shifts': night_s})
    return {
        'total_slots': total, 'filled_slots': filled, 'open_slots': open_,
        'fill_rate': f"{100*filled/total:.0f}%" if total else "0%",
        'days': days, 'staff_count': len(staff),
    }


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def process_two_week_file(uploaded_file) -> Tuple[bytes, Dict, str]:
    content = uploaded_file.read() if hasattr(uploaded_file, 'read') else uploaded_file
    wb = load_workbook(io.BytesIO(content), data_only=True)

    skip = {'HOL Staffing BLANK (2)','Staff List','Pasting Sheet',
            '|Shift| Tracker','Last (SHIFT) Tracker','6 Wk Counter',
            'Shift Names','Named Ranges','FYTD Counters'}
    sheets = [s for s in wb.sheetnames if s not in skip]
    if not sheets:
        raise ValueError("No schedule sheet found in the workbook.")

    # Use the most recently-dated sheet (last by order)
    sheet_name = sheets[-1]

    results, staff, layout = run_schedule(wb, sheet_name)
    out_wb  = build_output(wb, sheet_name, results, staff, layout)
    summary = _summary(results, staff)

    buf = io.BytesIO()
    out_wb.save(buf)
    return buf.getvalue(), summary, sheet_name
